"""
excel_store.py -- ทุกอย่างที่คุยกับไฟล์ DataWebapp.xlsx อยู่ในไฟล์นี้

ไฟล์ Excel เก็บ 2 sheet ที่เราใช้:
  - sheet "Mail"           : รายชื่อ user (อีเมล), Line, CostCenter, Zone
  - sheet "EditCheckSheet" : รายการขอแก้ไข check sheet ที่รออนุมัติ

หมายเหตุเรื่องการเขียนไฟล์:
  Excel ไม่ใช่ database จริง ถ้ามีคนสองคนเขียนพร้อมกันไฟล์จะพัง
  เราเลยใช้ FileLock ล็อกไฟล์ก่อนเขียน + เซฟลงไฟล์ .tmp ก่อนแล้วค่อย replace
  (ยกมาจากฟังก์ชัน edit_sheet_action_in_excel ของโค้ดเดิม)
"""

import logging
import os
from typing import Any

import openpyxl.utils
import pandas as pd
from filelock import FileLock, Timeout
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from app.config import settings

logger = logging.getLogger(__name__)


class ExcelBusyError(Exception):
    """มีคนอื่นกำลังเขียนไฟล์อยู่ -- ให้ผู้ใช้ลองใหม่"""


# ============================================================================
#  ส่วนอ่าน (Read)
# ============================================================================

def _read_sheet(sheet_name: str, **kwargs: Any) -> pd.DataFrame:
    """อ่าน sheet แล้ว strip ช่องว่างหน้า-หลังชื่อคอลัมน์ทิ้ง"""
    df = pd.read_excel(settings.EXCEL_FILE_PATH, sheet_name=sheet_name, **kwargs)
    df.columns = df.columns.str.strip()
    return df


def read_mail_sheet() -> pd.DataFrame:
    """อ่าน sheet Mail ทั้งหมด"""
    return _read_sheet(settings.EXCEL_MAIL_SHEET)


def read_edit_check_sheet() -> pd.DataFrame:
    """อ่าน sheet EditCheckSheet ทั้งหมด"""
    return _read_sheet(settings.EXCEL_EDIT_SHEET)


def load_users() -> dict[str, dict]:
    """
    แปลง sheet Mail -> dict ของ user

    แทนที่ฟังก์ชัน get_user_password() เดิม ที่เขียนลง global USERS,
    USER_LINE_MAP, USER_COSTCENTER_MAP แยกกัน 3 ตัว

    หมายเหตุ: คอลัมน์ชื่อ "User" แต่ข้อมูลข้างในเป็น "อีเมล"
    เราจึงใช้อีเมลเป็น key และเป็นสิ่งที่ผู้ใช้กรอกตอน login

    ผลลัพธ์หน้าตาแบบนี้:
        {
          "somchai.k@kubota.com": {
              "lines": ["CH", "MSC"],
              "costcenters": ["DY31010403"],
          },
          ...
        }

    user คนเดียวอาจมีหลายแถวใน Excel (คนละ Line) เลยต้อง append เข้า list
    """
    df = read_mail_sheet()
    users: dict[str, dict] = {}

    for _, row in df.iterrows():
        username = str(row.get("User", "")).strip().lower()
        # ข้ามแถวว่าง หรือแถวที่ pandas อ่านมาเป็น "nan"
        if not username or username == "nan":
            continue

        entry = users.setdefault(
            username,
            {"lines": [], "costcenters": []},
        )

        line = str(row.get("Line", "")).strip()
        if line and line != "nan" and line not in entry["lines"]:
            entry["lines"].append(line)

        costcenter = str(row.get("CostCenter", "")).strip()
        if costcenter and costcenter != "nan" and costcenter not in entry["costcenters"]:
            entry["costcenters"].append(costcenter)

    # admin เป็น user พิเศษที่ไม่ได้อยู่ใน Excel -- เห็นทุก Line ทุก CostCenter
    users["admin"] = {"lines": [], "costcenters": []}
    return users


def load_zone_line_map() -> dict[str, str]:
    """map:  Line -> Zone   (เช่น {"CH": "Final-B"})"""
    try:
        df = read_mail_sheet()
        return (
            df.dropna(subset=["Line", "Zone"])
            .set_index("Line")["Zone"]
            .to_dict()
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning("อ่าน zone_line_map ไม่ได้: %s", exc)
        return {}


def load_zone_costcenter_map() -> dict[str, str]:
    """map:  CostCenter -> Zone"""
    try:
        df = read_mail_sheet()
        return (
            df.dropna(subset=["CostCenter", "Zone"])
            .set_index("CostCenter")["Zone"]
            .to_dict()
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning("อ่าน zone_costcenter_map ไม่ได้: %s", exc)
        return {}


# ============================================================================
#  ส่วนเขียน (Write)  -- ต้องล็อกไฟล์เสมอ
# ============================================================================

# ชื่อคอลัมน์ -> ตัวอักษรคอลัมน์ใน Excel (ยกมาจาก column_map เดิม)
EDIT_SHEET_COLUMN_LETTERS = {
    "LD FI": "G",
    "FM FI": "H",
    "Asst-FinalB": "I",
}


def update_edit_check_sheet(row_index: int, action: str, column_name: str) -> str:
    """
    เขียนคำว่า Approve / Reject ลงช่องใน sheet EditCheckSheet

    row_index  = index ของแถวใน DataFrame (เริ่มที่ 0)
    action     = "approve" หรือ "reject"
    column_name= "LD FI" / "FM FI" / "Asst-FinalB"

    ทำไม row_index + 2?
      แถวที่ 1 ใน Excel คือหัวตาราง ส่วน DataFrame index เริ่มที่ 0
      ดังนั้น DataFrame แถวที่ 0 = Excel แถวที่ 2
    """
    col_letter = EDIT_SHEET_COLUMN_LETTERS.get(column_name)
    if not col_letter:
        raise ValueError(f"ไม่รู้จัก Column: {column_name}")

    status_text = "Approve" if action == "approve" else "Reject"
    lock = FileLock(settings.excel_lock_path, timeout=10)

    try:
        with lock:
            workbook = load_workbook(settings.EXCEL_FILE_PATH)
            sheet = workbook[settings.EXCEL_EDIT_SHEET]

            col_index = openpyxl.utils.column_index_from_string(col_letter)
            cell = sheet.cell(row=row_index + 2, column=col_index)
            cell.value = status_text
            cell.font = Font(name="SCG", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # เซฟลงไฟล์ชั่วคราวก่อน แล้วค่อยสลับทับไฟล์จริง
            # ถ้าไฟดับกลางทาง ไฟล์จริงจะยังไม่พัง
            tmp_path = settings.EXCEL_FILE_PATH + ".tmp"
            workbook.save(tmp_path)
            workbook.close()
            os.replace(tmp_path, settings.EXCEL_FILE_PATH)

        return status_text

    except Timeout as exc:
        raise ExcelBusyError("ระบบกำลังถูกใช้งานโดยผู้ใช้อื่น โปรดลองใหม่อีกครั้ง") from exc


def find_row_index_by_no(no: int) -> int | None:
    """หา index ของแถวใน EditCheckSheet จากคอลัมน์ 'No.'"""
    df = read_edit_check_sheet()
    matches = df.index[df["No."] == no].tolist()
    return matches[0] if matches else None
